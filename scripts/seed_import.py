"""Import historical seed data from Income_Expense_updated.xlsx.

Imports three sheets:
  - Expense - Expense: transaction-level expense history
  - Income - 表格 1: income transactions
  - Planning - Monthly: budget envelopes

Idempotent: uses source_ref = "historical:{sheet}:{row_index}" for dedup.
"""

import logging
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from db.models import Base, BudgetEnvelope, Category, Household, Transaction, User  # noqa: E402
from src.config import settings  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

XLSX_PATH = Path(__file__).resolve().parent.parent / "data" / "Income_Expense_updated.xlsx"

# Split method normalization: map spreadsheet values → canonical A/K/A/K
SPLIT_METHOD_MAP = {
    "A": "A",
    "K": "K",
    "A/K": "A/K",
    "A/K/M": "A/K",  # 3-way split with Karen → treat as A/K, use actual per-person amounts
    "A/M": "A",       # Andrew + Karen → Andrew only in 2-person model
    "Other": "A",     # edge case (2 rows), Andrew has the amount
}
# "M" (Karen-only) is intentionally excluded — skip those rows

# Category normalization: map spreadsheet names → canonical names
CATEGORY_MAP = {
    "Tax": "Other",  # 5 rows of Tax entries → map to Other
}


def _to_decimal(value: object) -> Decimal:
    """Convert a cell value to Decimal, treating False/None/empty as zero."""
    if value is None or value is False or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _extract_date(value: object) -> date | None:
    """Extract a date from a cell value."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _get_category_lookup(session: Session, household_id: str) -> dict[str, str]:
    """Build a name → id lookup for categories."""
    categories = session.execute(
        select(Category).where(Category.household_id == household_id)
    ).scalars().all()
    return {cat.name: cat.id for cat in categories}


def _get_user_lookup(session: Session, household_id: str) -> dict[str, str]:
    """Build a person_code → id lookup for users."""
    users = session.execute(
        select(User).where(User.household_id == household_id)
    ).scalars().all()
    return {user.person_code: user.id for user in users}


def _source_ref_exists(session: Session, household_id: str, source_ref: str) -> bool:
    """Check if a transaction with this source_ref already exists."""
    result = session.execute(
        select(Transaction.id).where(
            Transaction.household_id == household_id,
            Transaction.source_ref == source_ref,
        )
    ).scalar_one_or_none()
    return result is not None


def import_expense_sheet(
    session: Session,
    wb: openpyxl.Workbook,
    household_id: str,
    category_lookup: dict[str, str],
) -> dict[str, int]:
    """Import the 'Expense - Expense' sheet."""
    ws = wb["Expense - Expense"]
    stats = {"imported": 0, "skipped": 0, "failed": 0}

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Skip separator rows (date in col A, rest None) and blank rows
        if row[1] is None:
            continue

        source_ref = f"historical:expense:{row_idx}"

        # Idempotency check
        if _source_ref_exists(session, household_id, source_ref):
            stats["skipped"] += 1
            continue

        try:
            cash_date = _extract_date(row[1])  # col B
            if cash_date is None:
                stats["failed"] += 1
                continue

            description = str(row[4] or "").strip()  # col E
            if not description:
                stats["failed"] += 1
                continue

            raw_category = str(row[5] or "").strip()  # col F
            category_name = CATEGORY_MAP.get(raw_category, raw_category)

            category_id = category_lookup.get(category_name)
            if category_id is None:
                logger.warning(
                    "Row %d: unknown category '%s' (original: '%s'), skipping",
                    row_idx, category_name, raw_category,
                )
                stats["failed"] += 1
                continue

            raw_split = str(row[11] or "").strip()  # col L
            if raw_split == "M":
                # Karen-only transaction — not relevant to A/K household
                stats["skipped"] += 1
                continue

            split_method = SPLIT_METHOD_MAP.get(raw_split)
            if split_method is None:
                logger.warning("Row %d: unknown split method '%s', skipping", row_idx, raw_split)
                stats["failed"] += 1
                continue

            price = _to_decimal(row[9])  # col J = Price
            andrew_amount = _to_decimal(row[12])  # col M = Andrew
            kristy_amount = _to_decimal(row[13])  # col N = Kristy

            year = int(row[2]) if row[2] else cash_date.year   # col C
            month = int(row[3]) if row[3] else cash_date.month  # col D

            txn = Transaction(
                household_id=household_id,
                statement_id=None,
                cash_date=cash_date,
                accounting_period_year=year,
                accounting_period_month=month,
                accounting_period_is_override=False,
                description=description,
                normalized_description=description.lower().strip(),
                original_amount=price,
                original_currency="CAD",
                fx_rate=Decimal("1.0"),
                fx_rate_source="statement",
                cad_amount=price,
                category_id=category_id,
                split_method=split_method,
                andrew_amount=andrew_amount,
                kristy_amount=kristy_amount,
                classifier_confidence=None,
                classifier_source="historical",
                needs_review=False,
                is_manually_reviewed=True,
                source="historical_import",
                source_ref=source_ref,
            )
            session.add(txn)
            stats["imported"] += 1

        except Exception:
            logger.exception("Row %d: unexpected error", row_idx)
            stats["failed"] += 1

    return stats


def import_income_sheet(
    session: Session,
    wb: openpyxl.Workbook,
    household_id: str,
    category_lookup: dict[str, str],
) -> dict[str, int]:
    """Import the 'Income - 表格 1' sheet."""
    ws = wb["Income - 表格 1"]
    stats = {"imported": 0, "skipped": 0, "failed": 0}

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Skip separator rows and blank rows
        if row[1] is None:
            continue

        source_ref = f"historical:income:{row_idx}"

        if _source_ref_exists(session, household_id, source_ref):
            stats["skipped"] += 1
            continue

        try:
            cash_date = _extract_date(row[1])  # col B
            if cash_date is None:
                stats["failed"] += 1
                continue

            description = str(row[4] or "").strip()  # col E
            if not description:
                stats["failed"] += 1
                continue

            raw_category = str(row[5] or "").strip()  # col F
            category_name = CATEGORY_MAP.get(raw_category, raw_category)

            category_id = category_lookup.get(category_name)
            if category_id is None:
                logger.warning(
                    "Row %d: unknown income category '%s', skipping",
                    row_idx, category_name,
                )
                stats["failed"] += 1
                continue

            raw_split = str(row[7] or "").strip()  # col H
            split_method = SPLIT_METHOD_MAP.get(raw_split, raw_split)
            if split_method not in ("A", "K", "A/K"):
                logger.warning("Row %d: unknown split method '%s', skipping", row_idx, raw_split)
                stats["failed"] += 1
                continue

            price = _to_decimal(row[6])  # col G = Price
            andrew_amount = _to_decimal(row[8])  # col I = Andrew
            kristy_amount = _to_decimal(row[9])  # col J = Kristy

            year = int(row[2]) if row[2] else cash_date.year   # col C
            month = int(row[3]) if row[3] else cash_date.month  # col D

            txn = Transaction(
                household_id=household_id,
                statement_id=None,
                cash_date=cash_date,
                accounting_period_year=year,
                accounting_period_month=month,
                accounting_period_is_override=False,
                description=description,
                normalized_description=description.lower().strip(),
                original_amount=price,
                original_currency="CAD",
                fx_rate=Decimal("1.0"),
                fx_rate_source="statement",
                cad_amount=price,
                category_id=category_id,
                split_method=split_method,
                andrew_amount=andrew_amount,
                kristy_amount=kristy_amount,
                classifier_confidence=None,
                classifier_source="historical",
                needs_review=False,
                is_manually_reviewed=True,
                source="historical_import",
                source_ref=source_ref,
            )
            session.add(txn)
            stats["imported"] += 1

        except Exception:
            logger.exception("Row %d: unexpected error", row_idx)
            stats["failed"] += 1

    return stats


def import_planning_sheet(
    session: Session,
    wb: openpyxl.Workbook,
    household_id: str,
    category_lookup: dict[str, str],
    user_lookup: dict[str, str],
) -> dict[str, int]:
    """Import the 'Planning - Monthly' sheet as budget envelopes.

    The sheet has no year column, so we seed for both 2025 and 2026.
    """
    ws = wb["Planning - Monthly"]
    stats = {"imported": 0, "skipped": 0}

    # Rows 3-11 are expense categories (row 12 is blank, row 13 is Total)
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=11, values_only=True), start=3):
        raw_name = str(row[0] or "").strip()  # col A = category name
        if not raw_name:
            continue

        category_id = category_lookup.get(raw_name)
        if category_id is None:
            logger.warning("Planning row %d: unknown category '%s', skipping", row_idx, raw_name)
            continue

        andrew_amount = _to_decimal(row[1])  # col B
        kristy_amount = _to_decimal(row[2])  # col C

        for period_year in (2025, 2026):
            for person_code, amount in [("A", andrew_amount), ("K", kristy_amount)]:
                user_id = user_lookup[person_code]

                # Idempotency check
                existing = session.execute(
                    select(BudgetEnvelope).where(
                        BudgetEnvelope.category_id == category_id,
                        BudgetEnvelope.user_id == user_id,
                        BudgetEnvelope.period_year == period_year,
                    )
                ).scalar_one_or_none()

                if existing:
                    stats["skipped"] += 1
                    continue

                envelope = BudgetEnvelope(
                    household_id=household_id,
                    category_id=category_id,
                    user_id=user_id,
                    period_year=period_year,
                    amount_cad=amount,
                )
                session.add(envelope)
                stats["imported"] += 1

    return stats


def main() -> None:
    if not XLSX_PATH.exists():
        logger.error("Excel file not found: %s", XLSX_PATH)
        sys.exit(1)

    logger.info("Loading %s", XLSX_PATH)
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    engine = create_engine(settings.database_url, echo=False)
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        household = session.execute(
            select(Household).where(Household.id == settings.household_id)
        ).scalar_one_or_none()

        if household is None:
            logger.error(
                "Household '%s' not found. Run seed_categories.py first.",
                settings.household_id,
            )
            sys.exit(1)

        category_lookup = _get_category_lookup(session, household.id)
        user_lookup = _get_user_lookup(session, household.id)

        logger.info("Importing expense transactions...")
        expense_stats = import_expense_sheet(session, wb, household.id, category_lookup)

        logger.info("Importing income transactions...")
        income_stats = import_income_sheet(session, wb, household.id, category_lookup)

        logger.info("Importing budget envelopes...")
        planning_stats = import_planning_sheet(
            session, wb, household.id, category_lookup, user_lookup
        )

        session.commit()

    wb.close()

    logger.info("=" * 60)
    logger.info("Import complete")
    logger.info(
        "Expenses:  %d imported, %d skipped (duplicate), %d failed",
        expense_stats["imported"],
        expense_stats["skipped"],
        expense_stats["failed"],
    )
    logger.info(
        "Income:    %d imported, %d skipped (duplicate), %d failed",
        income_stats["imported"],
        income_stats["skipped"],
        income_stats["failed"],
    )
    logger.info(
        "Envelopes: %d imported, %d skipped (duplicate)",
        planning_stats["imported"],
        planning_stats["skipped"],
    )


if __name__ == "__main__":
    main()
